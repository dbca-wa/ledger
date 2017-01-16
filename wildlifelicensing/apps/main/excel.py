from collections import defaultdict, OrderedDict
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from django.utils import six
from django.utils.text import Truncator
from django.http import HttpResponse


def load_workbook_content(filename):
    return load_workbook(filename)


def get_named_ranges(wb):
    return [r.name for r in wb.get_named_ranges()]


def get_sheet_titles(wb):
    return [ws.title for ws in wb.worksheets]


def get_sheet(wb, title, case_insensitive=True):
    titles = get_sheet_titles(wb)
    for ws_title in titles:
        match = ws_title.lower() == title.lower() if case_insensitive else ws_title == title
        if match:
            return wb.get_sheet_by_name(ws_title)
    return None


def get_or_create_sheet(wb, title):
    ws = get_sheet(wb, title)
    if ws is None:
        ws = wb.create_sheet()
        ws.title = title
    return ws


def find_cell_by_value(ws, value):
    for row in ws.rows:
        for cell in row:
            if cell.value == value:
                return cell
    return None


def get_cell_neighbour(cell, direction='down'):
    if direction == 'down':
        return cell.offset(row=1)
    elif direction == 'up':
        return cell.offset(row=-1)
    elif direction == 'right':
        return cell.offset(column=1)
    elif direction == 'left':
        return cell.offset(column=-1)
    else:
        raise Exception("Invalid Direction: " + direction + ". Should be [down|up|right|left]")


def is_blank_value(value):
    return value is None or is_empty_string(value)


def is_empty_string(value):
    return isinstance(value, six.string_types) and len(value.strip()) == 0


def is_cell_blank(cell):
    return is_blank_value(cell.value)


def is_all_blanks(cells):
    return len([c for c in cells if not is_cell_blank(c)]) == 0


def strip(value):
    return value.strip() if isinstance(value, six.string_types) else value


def get_value_for_key(ws, key, direction='down'):
    result = None
    key_cell = find_cell_by_value(ws, key)
    if key_cell is not None:
        result = get_cell_neighbour(key_cell, direction).value
    return result


def write_values(ws, top_left_row, top_left_column, values, direction='right', font=None):
    top_cell = ws.cell(row=top_left_row, column=top_left_column)
    write_values_from_cell(top_cell, values, direction, font)
    return ws


def write_values_from_cell(top_cell, values, direction='right', font=None):
    cell = top_cell
    for value in values:
        if font is not None:
            cell.font = font
        cell.value = value
        cell = get_cell_neighbour(cell, direction)
    return top_cell


def append_column(ws, values):
    """
    Append a column in the given worksheet.
    :param ws:
    :param values:
    :return: the range string of the column (ex: "A1:A23")
    """
    # find first empty col
    top_cell = ws.cell(row=1, column=1)
    while not is_cell_blank(top_cell):
        top_cell = get_cell_neighbour(top_cell, 'right')
    # write values
    write_values_from_cell(top_cell=top_cell, values=values, direction='down')
    # build the range string
    top = top_cell.coordinate
    bottom = "{col}{row}".format(col=top_cell.column, row=(len(values)))
    return "{}:{}".format(top, bottom)


def create_list_validation(value, strict=True, allow_blank=True):
    if type(value) == list:
        formula = _build_list_formula(value)
    else:
        formula = value
    dv = DataValidation(type="list", formula1=formula, showErrorMessage=strict,
                        allow_blank=allow_blank)
    if strict:
        dv.promptTitle = 'Strict Selection'
        dv.prompt = 'You must select a value from the list'
    else:
        dv.promptTitle = 'Proposed Selection'
        dv.prompt = 'You may select a value from the list or enter your own'
    if allow_blank:
        dv.promptTitle += " (blank allowed)"
    return dv


def _build_list_formula(values):
    """
    :return: a quoted (") comma separated string: '1,2,3' that can be applied on data validation formula
    WARNING: there's a 255 characters limit for a formula literal in Excel.
    """
    csv_values = Truncator(",".join(values)).chars(255)
    return '"{}"'.format(csv_values)


class TableData:
    """
    Parse a square portion of a spreadsheet.
    It supports two modes: normal or transpose where columns are rows and rows are columns
    The top left (row,col) starting at (1,1) is given as parameter.
    The first row represent the column_headers
    The column header parsing stop at the first empty cell on the first row
    The row parsing stops at the first blank row.
    :param transpose: if true the the table is a transposed one. Columns are rows and rows are columns
    """

    def __init__(self, worksheet, top_left_row=1, top_left_column=1, nb_cols=None, nb_rows=None, transpose=False):
        self.worksheet = worksheet
        self.top_left_row = top_left_row
        self.top_left_column = top_left_column
        self.transpose = transpose
        self.top_left_cell = worksheet.cell(row=top_left_row, column=top_left_column)
        self.column_headers = self._parse_column_headers()
        self.rows = self._parse_rows()

    def by_columns(self):
        """
        :return: [(col_header1, [row1, row2,..]), (col_header2, [row1, row2,..]), ...]
        """
        result = defaultdict(list)
        for row in self.rows:
            for column, value in zip(self.column_headers, row):
                result[column].append(value)
        return result.items()

    def by_rows(self):
        """
        :return: [[(col_header1, row1), (col_header2, row1)...],  [(col_header1, row2), (col_header2, row2), ...]]
        """
        return [zip(self.column_headers, row) for row in self.rows]

    def rows_by_col_header_it(self):
        """
        A row iterator
        :return: a OrderedDict like:
                {
                    ....
                    'col_header1': row_n_col1
                    'col_header2': row_n-col2
                    .......
                },
        """
        for row in self.rows:
            data = OrderedDict()
            for i, col_header in enumerate(self.column_headers):
                if col_header not in data:
                    data[col_header] = row[i]
                else:
                    # if they are two columns with the same header we store it with with a appended _i
                    count = 1
                    key = col_header + '_' + str(count)
                    while key in data:
                        count += 1
                        key = col_header + '_' + str(count)
                    data[key] = row[i]
            yield data

    def rows_by_col_letter_it(self):
        """
        A row iterator
        :return: a dict like:
         {'A': value1,
          'B': value2,
          .....
        }
        """
        for row in self.rows:
            data = {}
            for i, value in enumerate(row):
                data[get_column_letter(i + 1)] = value
            yield data

    def _parse_column_headers(self):
        headers = []
        cell = self.top_left_cell
        while not is_cell_blank(cell):
            headers.append(strip(cell.value))
            direction = 'down' if self.transpose else 'right'
            cell = get_cell_neighbour(cell, direction)
        return headers

    def _parse_rows(self):
        rows = []
        start_row = self.top_left_row if self.transpose else self.top_left_row + 1
        start_col = self.top_left_column + 1 if self.transpose else self.top_left_column
        blank_row = False
        row_index = start_row
        col_index = start_col
        while not blank_row:
            row_cells = []
            for i in range(0, len(self.column_headers)):
                cell = self.worksheet.cell(row=row_index, column=col_index)
                row_cells.append(cell)
                if self.transpose:
                    row_index += 1
                else:
                    col_index += 1
            if is_all_blanks(row_cells):
                blank_row = True
            else:
                rows.append([strip(c.value) for c in row_cells])
                if self.transpose:
                    col_index += 1
                    row_index = start_row
                else:
                    row_index += 1
                    col_index = start_col
        return rows

    def _get_row_cells(self, row_index):
        start = self.top_left_column - 1
        end = start + len(self.column_headers)

        return self.worksheet.rows[row_index][start:end]


class ExcelFileResponse(HttpResponse):
    def __init__(self, content, file_name=None):
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        content_disposition = 'attachment;'

        if file_name is not None:
            if not file_name.lower().endswith('.xlsx'):
                file_name += '.xlsx'
            content_disposition += ' filename=' + file_name

        super(ExcelFileResponse, self).__init__(content, content_type=content_type)
        self['Content-Disposition'] = content_disposition


class WorkbookResponse(ExcelFileResponse):
    def __init__(self, wb, file_name=None):
        super(WorkbookResponse, self).__init__([], file_name=file_name)
        wb.save(self)
