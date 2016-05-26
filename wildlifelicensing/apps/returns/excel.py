from openpyxl import load_workbook


def load_workbook_headings(filename):
    workbook = load_workbook(filename=filename, read_only=True)
    worksheet = workbook.worksheets[0]

    # worksheet.rows is a generator, so to access just the top row, iterate just the first item and return
    for row in worksheet.rows:
        return [cell.value for cell in row]


def load_workbook(filename):
    workbook = load_workbook(filename=filename, read_only=True)
