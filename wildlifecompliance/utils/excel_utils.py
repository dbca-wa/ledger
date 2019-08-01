from __future__ import unicode_literals

from django.conf import settings
from collections import OrderedDict
from wildlifecompliance.components.applications.models import Application, ExcelApplication, ExcelActivityType
from wildlifecompliance.components.licences.models import DefaultActivityType, WildlifeLicence, WildlifeLicenceClass, WildlifeLicenceActivity
from wildlifecompliance.components.organisations.models import Organisation
from ledger.accounts.models import OrganisationAddress
from django.contrib.postgres.fields.jsonb import JSONField
from wildlifecompliance.utils import SearchUtils, search_multiple_keys
from wildlifecompliance.components.licences.models import DefaultActivity
from ledger.accounts.models import EmailUser

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell, xl_cell_to_rowcol, xl_col_to_name
import xlrd#, xlwt
import openpyxl
from openpyxl.styles import Protection
import os
import re
import shutil

import logging
logger = logging.getLogger(__name__)


import json
from datetime import datetime
from django.db import models


APP_SHEET_NAME = 'Applications'
META_SHEET_NAME = 'Meta'
SYSTEM = 'System'
FIRST_COL = 'First Col'
LAST_COL = 'Last Col'

NSYS_COLS = 6 # number of system columns (cols --> 0 to 5)
LODGEMENT_NUMBER = 'lodgement_number'
APPLICATION_ID = 'application_id'
LICENCE_NUMBER = 'licence_number'
APPLICANT = 'applicant'
APPLICANT_TYPE = 'applicant_type'
APPLICANT_ID = 'applicant_id'

PURPOSE = 'purpose'
ADDITIONAL_INFO = 'additional_info'
STANDARD_ADVANCED = 'standard_advanced'
COVER_PROCESSED = 'cover_processed'
COVER_PROCESSED_DATE = 'cover_processed_date'
COVER_PROCESSED_BY = 'cover_processed_by'
CONDITIONS = 'conditions'
ISSUE_DATE = 'issue_date'
START_DATE = 'start_date'
EXPIRY_DATE = 'expiry_date'
TO_BE_ISSUED = 'to_be_issued'
PROCESSED = 'processed'

APPLICANT_TYPE_ORGANISATION = 'ORG'
APPLICANT_TYPE_PROXY = 'PRX'
APPLICANT_TYPE_SUBMITTER = 'SUB'


class ExcelWriter():
    def __init__(self):
        self.cur_datetime = datetime.now()
        if not ExcelApplication.objects.exists():
            # hack to allow templates to run when no ExcelApplication exists yet - need the System header columns from the excel_app object (cols_output)
            app = Application.objects.all().last()
            self.create_excel_model(app.licence_category, use_id=app.id)

    def update_workbooks(self):
        for licence_category in WildlifeLicenceClass.objects.all():
            filename = '{}.xlsx'.format(self.replace_special_chars(licence_category.short_name))
            self.archive_workbook(filename)
            self.read_workbook(settings.EXCEL_OUTPUT_PATH + '/' + filename, licence_category.short_name)
            logger.info('Completed: ' + licence_category.short_name + ' - ' + settings.EXCEL_OUTPUT_PATH + '/' + filename)

    def archive_workbook(self, filename):
        src_file = settings.EXCEL_OUTPUT_PATH + '/' + filename
        if os.path.isfile(src_file):
            archive_dir = '{}/archive/{}'.format(settings.EXCEL_OUTPUT_PATH, self.cur_datetime.strftime('%Y%m%dT%H%M%S'))
            dest_file = archive_dir + '/' + filename
            if not os.path.exists(archive_dir):
                os.makedirs(archive_dir)
            shutil.copyfile(src_file, dest_file)


    def set_formats(self, workbook):
        self.bold = workbook.add_format({'bold': True})
        self.bold_unlocked = workbook.add_format({'bold': True, 'locked': False})
        self.unlocked = workbook.add_format({'locked': False})
        self.locked = workbook.add_format({'locked': True})
        self.wrap = workbook.add_format({'text_wrap': True})
        self.unlocked_wrap = workbook.add_format({'text_wrap': True, 'locked': False})
        self.integer = workbook.add_format({'num_format': '0', 'align': 'center'})

    def replace_special_chars(self, input_str, new_char='_'):
        return re.sub('[^A-Za-z0-9]+', new_char, input_str).strip('_').lower()

    def get_purposes(self, licence_class_short_name):
        """
            activity_type --> purpose

            for licence_class in DefaultActivityType.objects.filter(licence_class_id=13):
                #print '{} {}'.format(licence_class.licence_class.id, licence_class.licence_class.name)
                for activity_type in DefaultActivity.objects.filter(activity_type_id=licence_class.activity_type_id):
                    print '    {}'.format(activity_type.activity.name, activity_type.activity.short_name)
            ____________________

            DefaultActivityType.objects.filter(licence_class__short_name='Flora Other Purpose').values_list('licence_class__activity_type__activity__name', flat=True).distinct()
        """
        activity_type =  DefaultActivityType.objects.filter(licence_class__short_name=licence_class_short_name).order_by('licence_class__activity_type__activity__name')
        return activity_type.values_list('licence_class__activity_type__activity__name', flat=True).distinct()

    def get_licence_type(self, activity_type_name):
        """
        activity_type name -- purpose name--> 'Importing Fauna (Non-Commercial)'
        """
        return DefaultActivityType.objects.filter(licence_class__activity_type__activity__name=activity_type_name).distinct('licence_class')[0].licence_class.name

    def get_index(self, values_list, name):
        indices = [i for i, s in enumerate(values_list) if name in s]
        return indices[0] if indices else None

    def get_activity_type_sys_questions(self, activity_name):
        """
        Looks up the activity type schema and return all questions (marked isEditable) that need to be added to the Excel WB.
        Allows us to know the block size for each activity type in the WB (start_col, end_col)
        """
        ordered_dict=OrderedDict([])

        schema = WildlifeLicenceActivity.objects.get(name=activity_name).schema
        res = search_multiple_keys(schema, 'isEditable', ['name', 'label'])
        [ordered_dict.update([(i['name'],i['label'])]) for i in res]

        return ordered_dict

    def get_tab_index(self, qs_activity_type):
        application = qs_activity_type[0].application
        activity_name = qs_activity_type[0].name # 'Fauna Other - Importing'
        return application.data[0][activity_name][0].keys()[0].split('_')[1]

    def get_activity_type_sys_answers(self, qs_activity_type, activity_name):
        """
        Looks up the activity type return all answers for question marked isEditable that need to be added to the Excel WB.
        """
        ordered_dict=OrderedDict([])
        questions = self.get_activity_type_sys_questions(activity_name)
        for k,v in questions.iteritems():
            # k - section name
            # v - question
            if qs_activity_type:
                # must append tab index to 'section name'
                k = k + '_' + self.get_tab_index(qs_activity_type) #, activity_name)
                s = SearchUtils(qs_activity_type[0].application)
                answer = s.search_value(k)
                #ordered_dict.update(OrderedDict([(v,answer)]))
                ordered_dict.update(OrderedDict([(k,answer)]))
            else:
                # this part for the Excel column  headers
                #ordered_dict.update(OrderedDict([(v,None)]))
                ordered_dict.update(OrderedDict([(k,None)]))

        return ordered_dict

    def create_excel_model(self, licence_category, cur_app_ids=[], use_id=None):
        """
        from wildlifecompliance.utils.excel_utils import write_excel_model
        write_excel_model('Fauna Other Purpose')
        """

        if use_id:
            # get a filterset with single application
            applications = Application.objects.filter(id=Application.objects.all().last().id)
        else:
            applications = Application.objects.filter(licence_category=licence_category).exclude(processing_status=Application.PROCESSING_STATUS_DRAFT[0]).exclude(id__in=cur_app_ids)

        new_excel_apps = []
        for application in applications.order_by('id'):
            excel_app, created = ExcelApplication.objects.get_or_create(application=application)
            new_excel_apps.append(excel_app)

            activities = self.get_purposes(application.licence_type_data['short_name']).values_list('activity_type__short_name', flat=True)
            for activity_type in application.licence_type_data['activity_type']:
                if activity_type['short_name'] in list(activities):
                    excel_activity_type, created = ExcelActivityType.objects.get_or_create(
                        excel_app=excel_app,
                        activity_name=activity_type['activity'][0]['name'],
                        name=activity_type['name'],
                        short_name=activity_type['short_name']
                    )

        return new_excel_apps

    def create_workbook_template(self, filename, licence_category='Fauna Other Purpose'):
        """
        Creates a blank template with purposes and column headings only
        """
        meta = OrderedDict()
        if os.path.isfile(filename):
            logger.warn('File already exists {}'.format(filename))
            return None

        wb = xlsxwriter.Workbook(filename)
        ws = wb.add_worksheet(APP_SHEET_NAME)
        self.set_formats(wb)

        row_num = 0
        col_num = 0
        cell_dict = {}
        cell_start = xl_rowcol_to_cell(row_num, col_num, row_abs=True, col_abs=True)
        sys_cols = ExcelApplication.objects.all().last().cols_output.keys()
        for col_name in sys_cols:
            ws.write(row_num, col_num, col_name, self.bold_unlocked)
            col_num += 1
        cell_end = xl_rowcol_to_cell(row_num, col_num-1, row_abs=True, col_abs=True)
        cell_dict.update({SYSTEM: [cell_start, cell_end]})

        activity_name_list = self.get_purposes(licence_category)
        for activity_name in activity_name_list:
            #cols = self.cols_output(None, 'Importing Fauna (Non-Commercial)')
            activity_type_cols = self.cols_output(None, activity_name).keys()
            ws.write(row_num, col_num, '', self.bold); col_num += 1
            cell_start = xl_rowcol_to_cell(row_num, col_num, row_abs=True, col_abs=True)
            for col_name in activity_type_cols:
                ws.write(row_num, col_num, col_name, self.bold_unlocked)
                col_num += 1
            cell_end = xl_rowcol_to_cell(row_num, col_num-1, row_abs=True, col_abs=True)
            cell_dict.update({activity_name: [cell_start, cell_end]})

        self.write_sheet_meta(wb, cell_dict, activity_name_list)
        wb.close()

    def read_workbook(self, input_filename, licence_category='Fauna Other Purpose'):
        """
        Read the contents of input_filename, create licences if to_be_processed='Y' and append new applications to the workbook
        """
        meta = OrderedDict()
        if not os.path.isfile(input_filename):
            logger.warn('Cannot find file {}. Creating ...'.format(input_filename))
            self.create_workbook_template(input_filename, licence_category)

        wb = xlrd.open_workbook(input_filename)
        sh = wb.sheet_by_name(APP_SHEET_NAME)
        sh_meta = wb.sheet_by_name(META_SHEET_NAME)

        # Read Meta
        number_of_rows = sh_meta.nrows
        hdr = sh_meta.row_values(0)
        for row in range(1, number_of_rows):
            row_values = sh_meta.row_values(row)
            purpose = row_values[0]
            meta.update([(purpose, {})])
            for i in zip(hdr, row_values)[1:]:
                meta[purpose].update( {i[0]: i[1]} )

        # Read Application Data
        excel_data = {}
        number_of_rows = sh.nrows
        hdr = sh.row_values(0)
        for row in range(1, number_of_rows):
            row_values = sh.row_values(row)
            lodgement_number = row_values[hdr.index(LODGEMENT_NUMBER)]
            application_id = int(row_values[hdr.index(APPLICATION_ID)])
            licence_number = row_values[hdr.index(LICENCE_NUMBER)]
            applicant = row_values[hdr.index(APPLICANT)]
            applicant_type = row_values[hdr.index(APPLICANT_TYPE)]
            applicant_id = int(row_values[hdr.index(APPLICANT_ID)])
            application = Application.objects.get(lodgement_number=lodgement_number)

            for purpose in meta.keys():
                if purpose != SYSTEM:
                    try:
                        idx_start = int(meta[purpose][FIRST_COL])
                        idx_end = int(meta[purpose][LAST_COL])
                        purpose_row = row_values[idx_start:idx_end]
                        hdr_row = hdr[idx_start:idx_end]


                        idx_to_be_issued = self.get_index(hdr_row, TO_BE_ISSUED)
                        to_be_issued = purpose_row[idx_to_be_issued]
                        if to_be_issued in ['y', 'Y'] and not licence_number:
                            # create licence, if not already created
                            # check if user already has a licence. if so, re-use licence_number and update the licence_sequence
                            if application.licences.all().count() > 0:
                                licence_number = application.licences.all().first().reference
                            else:
                                licence_number = self.create_licence(application, purpose, licence_category, applicant_type, applicant_id).reference
                            row_values[hdr.index(LICENCE_NUMBER)] = licence_number
                    except ValueError, e:
                        logger.error('Cannot find activity_type {} in Excel header./n{}'.format(purpose, e))
#                    except Exception, e:
#                        import ipdb; ipdb.set_trace()

            excel_data.update({lodgement_number: row_values})

        #wb.release_resources()
        #del wb

        # Re-output Application Data with licence_numbers
        #wb = xlsxwriter.Workbook(input_filename)
        #self.set_formats(wb_out)
        #ws_out = wb_out.add_worksheet(APP_SHEET_NAME)

        #ws = wb.get_worksheet_by_name(APP_SHEET_NAME)
        #sh_meta = wb.sheet_by_name(META_SHEET_NAME)

        wb = openpyxl.load_workbook(input_filename)
        ws = wb.get_sheet_by_name(APP_SHEET_NAME)
        ws.protection.sheet = True

        row_num = 0
        col_num = 0
        #ws.write_row(row_num, col_num, hdr, self.bold); row_num += 1
        #import ipdb; ipdb.set_trace()
        self.write_row(row_num, col_num, hdr, ws, is_header=True); row_num += 1
        for k,v in excel_data.iteritems():
            #ws.write_row(row_num, col_num, v)
            self.write_row(row_num, col_num, v, ws)
            row_num += 1

        # Append new applications to output
        #cur_app_ids = [int(v[1]) for k,v in excel_data.iteritems()] # existing app id's
        #new_app_ids = Application.objects.exclude(processing_status='draft').exclude(id__in=cur_app_ids)
        self.write_new_app_data(excel_data, meta, licence_category, ws, row_num)

        #wb_out.close()
        try:
            wb.save(input_filename)
        except IOError, e:
            raise Exception("Cannot write to file {}. File is already open. Please close the file first".format(input_filename))


    def write_row(self, row_num, col_num, values, worksheet, is_header=False): # openpyxl helper function
        """ Openpyxl helper function. Writes values as a row of data. If values is a single value, writes to a single cell """
        if not isinstance(values, list):
            values = [values] if values else ['']

        for col, val in enumerate(values, start=col_num):
            cell = worksheet.cell(row=row_num+1, column=col+1, value=val) # openpyxl count rows and cols from 1
            if not is_header and col > NSYS_COLS:
                # don't protect data cells
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)

    def write_new_app_data(self, excel_data, meta, licence_category, worksheet, next_row):

        cur_app_ids = [int(v[1]) for k,v in excel_data.iteritems()] # existing app id's
        new_excel_apps = self.create_excel_model(licence_category, cur_app_ids=cur_app_ids)

        row_num = next_row
        for excel_app in new_excel_apps:

            # System data
            col_num = int(meta[SYSTEM][FIRST_COL])
            for k,v in excel_app.cols_output.iteritems():
                #worksheet.write(row_num, col_num, v, self.locked)
                self.write_row(row_num, col_num, v, worksheet)
                col_num += 1

            # Application data
            for purpose in meta.keys():

                if purpose != SYSTEM:
                    activity_type = excel_app.excel_activity_types.filter(activity_name=purpose)
                    activity_type_cols = self.cols_output(activity_type, purpose)

                    col_num = int(meta[purpose][FIRST_COL])# + 1
                    if activity_type.exists():
                        for k,v in activity_type_cols.iteritems():
                            #ws.write('B1', 'Here is\nsome long text\nthat\nwe wrap',      wrap)
                            #worksheet.write(row_num, col_num, v, self.unlocked_wrap)
                            self.write_row(row_num, col_num, v, worksheet)
                            col_num += 1
                    else:
                        # create a blank activity_type bilock
                        for _ in activity_type_cols.keys():
                            #worksheet.write(row_num, col_num, '', self.unlocked)
                            self.write_row(row_num, col_num, '', worksheet)
                            col_num += 1

            row_num += 1

#    def cols_system(self, qs_activity_type, activity_name):
#        """ qs_excel_app --> ExcelApplication """
#        return OrderedDict([
#            ('lodgement_number', qs_activity_type[0].excel_app.lodgement_number if qs_activity_type else None),
#            ('application_id', qs_activity_type[0].excel_app.application.id if qs_activity_type else None),
#            ('licence_number', qs_activity_type[0].excel_app.licence_number if qs_activity_type else None),
#            ('applicant', qs_activity_type[0].excel_app.applicant_details if qs_activity_type else None),
#            ('applicant_type', qs_activity_type[0].excel_app.applicant_type if qs_activity_type else None),
#            ('applicant_id', qs_activity_type[0].excel_app.applicant_id if qs_activity_type else None),
#        ])

    def cols_common(self, qs_activity_type, activity_name, code):
        #code = activity_name[:2].lower()
        ordered_dict = OrderedDict([
            ('{}_{}'.format(code, PURPOSE), None),
            ('{}_{}'.format(code, ADDITIONAL_INFO), None),
            ('{}_{}'.format(code, STANDARD_ADVANCED), None),
            ('{}_{}'.format(code, COVER_PROCESSED), None),
            ('{}_{}'.format(code, COVER_PROCESSED_DATE), None),
            ('{}_{}'.format(code, COVER_PROCESSED_BY), None),
            ('{}_{}'.format(code, CONDITIONS), qs_activity_type[0].conditions if qs_activity_type else None),
            ('{}_{}'.format(code, ISSUE_DATE), qs_activity_type[0].issue_date if qs_activity_type else None),
            ('{}_{}'.format(code, START_DATE), qs_activity_type[0].start_date if qs_activity_type else None),
            ('{}_{}'.format(code, EXPIRY_DATE), qs_activity_type[0].expiry_date if qs_activity_type else None),
            ('{}_{}'.format(code, TO_BE_ISSUED), qs_activity_type[0].issued if qs_activity_type else None),
            ('{}_{}'.format(code, PROCESSED), qs_activity_type[0].processed if qs_activity_type else None),
        ])
        return ordered_dict

    def cols_output(self, qs_activity_type, activity_name):
        """
        excel_app = ExcelApplication.objects.all().last()
        activity_type = excel_app.excel_activity_types.filter(activity_name='Importing Fauna (Non-Commercial)')[0]
        cols_output(activity_type, 'Importing', 'Importing Fauna (Non-Commercial)')
        """
        code = WildlifeLicenceActivity.objects.get(name=activity_name).code
        ordered_dict = OrderedDict([
            ('{}'.format(activity_name), None),
            ('{}'.format(code), None),
        ])
        ordered_dict.update(self.get_activity_type_sys_answers(qs_activity_type, activity_name))
        ordered_dict.update(self.cols_common(qs_activity_type, activity_name, code))
        return ordered_dict


    #def create_licence(self, application, activity_name, applicant_type, applicant_id):
    def create_licence(self, application, activity_name, licence_category, applicant_type, applicant_id):
        """ activity_name='Importing Fauna (Non-Commercial)' """
        licence = None
        activity=WildlifeLicenceActivity.objects.get(name=activity_name)
        licence_class = WildlifeLicenceClass.objects.get(short_name=licence_category)
        if applicant_type == APPLICANT_TYPE_ORGANISATION:
            qs_licence = WildlifeLicence.objects.filter(org_applicant_id=applicant_id, licence_class=licence_class)
            if qs_licence.exists():
                # use existing licence, just increment sequence_number
                licence = WildlifeLicence.objects.create(
                    licence_number=qs_licence.last().licence_number,
                    licence_sequence=qs_licence.last().licence_sequence + 1,
                    current_application=application,
                    org_applicant_id=applicant_id,
                    submitter=application.submitter,
                    licence_type=activity,
                    licence_class=licence_class
                )
            else:
                licence = WildlifeLicence.objects.create(current_application=application,org_applicant_id=applicant_id,submitter=application.submitter,
                              licence_type=activity,licence_class=licence_class)

        elif applicant_type == APPLICANT_TYPE_PROXY:
            qs_licence = WildlifeLicence.objects.filter(proxy_applicant_id=applicant_id, licence_class=licence_class)
            if qs_licence.exists():
                # use existing licence, just increment sequence_number
                licence = WildlifeLicence.objects.create(
                    licence_number=qs_licence.last().licence_number,
                    licence_sequence=qs_licence.last().licence_sequence + 1,
                    current_application=application,
                    proxy_applicant_id=applicant_id,
                    submitter=application.submitter,
                    licence_type=activity,
                    licence_class=licence_class
                )
            else:
                licence = WildlifeLicence.objects.create(current_application=application, proxy_applicant_id=applicant_id, submitter=application.submitter,
                              licence_type=activity, licence_class=licence_class)

        #elif applicant_type == APPLICANT_TYPE_SUBMITTER:
        else: # assume applicant is the submitter
            qs_licence = WildlifeLicence.objects.filter(submitter_id=applicant_id, org_applicant__isnull=True, proxy_applicant__isnull=True, licence_class=licence_class)
            if qs_licence.exists():
                # use existing licence, just increment sequence_number
                licence = WildlifeLicence.objects.create(
                    licence_number=qs_licence.last().licence_number,
                    licence_sequence=qs_licence.last().licence_sequence + 1,
                    current_application=application,
                    #proxy_applicant_id=applicant_id,
                    submitter_id=applicant_id,
                    licence_type=activity,
                     licence_class=licence_class
                )
            else:
                licence = WildlifeLicence.objects.create(current_application=application, submitter_id=applicant_id, licence_type=activity, licence_class=licence_class)

        return licence

    def write_sheet_meta(self, workbook, cell_dict, activity_name_list):
        ws_meta = workbook.add_worksheet(META_SHEET_NAME)
        #ws_meta = workbook.sheet_by_name('Meta')

        row_num = 0
        col_num = 0
        workbook.define_name('{0}!{1}'.format(APP_SHEET_NAME, SYSTEM), '={0}!{1}:{2}'.format(APP_SHEET_NAME, cell_dict[SYSTEM][0], cell_dict[SYSTEM][1]))
        # Hdr
        ws_meta.write(row_num, col_num, 'Purpose')
        ws_meta.write(row_num, col_num+1, FIRST_COL)
        ws_meta.write(row_num, col_num+2, LAST_COL)

        # System
        ws_meta.write(row_num+1, col_num, SYSTEM)                                         # Purpose
        ws_meta.write(row_num+1, col_num+1, xl_cell_to_rowcol(cell_dict[SYSTEM][0])[1])   # First Col
        ws_meta.write(row_num+1, col_num+2, xl_cell_to_rowcol(cell_dict[SYSTEM][1])[1])   # Last Col

        # Purposes
        width = 20
        for row_num, activity_name in enumerate(activity_name_list, start=2):
            workbook.define_name('{0}!{1}'.format(APP_SHEET_NAME, self.replace_special_chars(activity_name)), '={0}!{1}:{2}'.format(APP_SHEET_NAME, cell_dict[activity_name][0], cell_dict[activity_name][1]))
            ws_meta.write(row_num, col_num, activity_name)                                          # Purpose
            ws_meta.write(row_num, col_num+1, xl_cell_to_rowcol(cell_dict[activity_name][0])[1])    # First Col
            ws_meta.write(row_num, col_num+2, xl_cell_to_rowcol(cell_dict[activity_name][1])[1])    # Last Col
            width = len(activity_name) if len(activity_name) > width else width

        #  lock  the Meta Sheet (mostly)
        ws_meta.set_column('D:G', 13, self.integer )
        ws_meta.set_column('A:Z', None, self.locked)
        ws_meta.protect()
        ws_meta.set_column(0, 0, width)



def read_codes():
    code_list = [
         ('Bioprospecting - Flora or Fauna', 'BIO'),
         ('Exporting Flora (Non-Commercial)', 'EFLN'),
         ('Taking Flora for Biological Survey', 'TFLB'),
         ('Taking Flora from Crown Land (Non-Commercial)', 'TFLN'),
         ('Exporting Sandalwood (Commercial)', 'ES'),
         ('Dealing in Sandalwood', 'DS'),
         ('Sandalwood Processing Establishment', 'SPE'),
         ('Private Property Supplier of Sandalwood', 'SS'),
         ('Taking Sandalwood from Crown Land (Commercial)', 'TSC'),
         ('Exporting Flora (Commercial)', 'EFLC'),
         ('Dealing in Flora', 'DFL'),
         ('Flora Processing Establishment', 'FLPE'),
         ('Private Property Supplier', 'PPS'),
         ('Taking Flora from Crown Land (Commercial)', 'TFLC'),
         ('Exporting Fauna (Non-Commercial)', 'EFA'),
         ('Importing Fauna (Non-Commercial)', 'IFA'),
         ('Possessing Fauna, Pet Keeping', 'PFAP'),
         ('Possessing Fauna for Research', 'PFAR'),
         ('Possessing Derelict Fauna', 'PDFA'),
         ('Fauna Interaction (Non-Commercial)', 'FAIN'),
         ('Salvage and Translocation of Fauna', 'STFA'),
         ('Fauna Biological Survey', 'FABS'),
         ('Fauna Causing Damage', 'FACD'),
         ('Taking Dangerous Fauna  ', 'TDFA'),
         ('Scientific / Education / Research', 'SCER'),
         ('Exporting Kangaroo Products', 'EKP'),
         ('Dealing in Kangaroo Products', 'DKP'),
         ('Processing Kangaroos', 'PK'),
         ('Possessing Kangaroos (Including in chiller facilities)', 'POK'),
         ('Taking Fauna (Kangaroos)', 'TFAK'),
         ('Exporting Fauna (Commercial)', 'EFAC'),
         ('Importing Fauna (Commercial)', 'IFAC'),
         ('Pet Dealer', 'PDE'),
         ('Dealer', 'DE'),
         ('Processing Fauna (e.g. crocodiles)', 'PFA'),
         ('Commercial Display of Fauna (includes Wildlife Parks)', 'DFA'),
         ('Feeding of Fauna for Commercial Purposes', 'FFA'),
         ('Fauna Interaction (commercial)', 'FAIC'),
         ('Taking Live Fauna for Commercial Purpose', 'TFAC'),
         ('Taking Fauna (destroying)', 'TFAD')
    ]

    for j,i in enumerate(code_list):
        try:
            a = WildlifeLicenceActivity.objects.get(name=i[0].strip())
            a.code = i[1]
            a.save()
        except Exception, e:
            print j, i[0], i[1], e
